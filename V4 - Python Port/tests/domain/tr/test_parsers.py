"""
Parser tests use in-memory openpyxl workbooks and a stub resolver so they
never need live domain services or a real Excel file.
"""
from datetime import date, datetime
from typing import Any

import openpyxl
import pytest
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

from infrastructure.database import Base
from slim.domain.analysis.analysis_service import AnalysisService
from slim.domain.chemical.chemical import Chemical
from slim.domain.chemical.chemical_service import ChemicalService
from slim.domain.customer.customer import Customer
from slim.domain.customer.customer_service import CustomerService
from slim.domain.element.element_repository import _ElementRow
from slim.domain.element.element_service import ElementService
from slim.domain.tr.enums import ProcessingTime, RequestType
from slim.domain.tr.tr_sample_form_parser import (
    SampleRangeBounds,
    TRSampleFormParser,
    _parse_element_symbols,
)
from slim.domain.tr.tr_submission_form_parser import (
    TRSubmissionFormParser,
    _extract_service_date,
    _parse_emails,
)
from slim.domain.tr.tr_submission import SENTINEL_DATE

# ---------------------------------------------------------------------------
# Stub resolver — returns seeded fixtures regardless of input
# ---------------------------------------------------------------------------

_STUB_CUSTOMER = Customer(
    id=1,
    name="Acme Corp",
    street_address="123 Main St",
    city="Springfield",
    state="IL",
    postal_code="62701",
    country="USA",
)
_STUB_CHEMICAL = Chemical(
    id=7,
    name="Hydrochloric Acid",
    metals_prep="Dilute and Shoot",
    silicon_prep="N/A",
    ions_prep="Dilute and Shoot",
)


class _StubResolver:
    def resolve_customer(self, name: str, address: str, address2: str) -> Customer:
        return _STUB_CUSTOMER

    def resolve_chemical(self, form_name: str) -> Chemical:
        return _STUB_CHEMICAL


# ---------------------------------------------------------------------------
# Minimal service fixtures (using in-memory SQLite)
# ---------------------------------------------------------------------------


@pytest.fixture
def session() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        # One element so AddElement tests can resolve a symbol
        s.add(_ElementRow(id=14, element_symbol="Si", element_name="Silicon"))
        s.commit()
        # Create invoice_default_contacts so the parser query doesn't fail
        s.execute(
            text(
                "CREATE TABLE invoice_default_contacts "
                "(contact_email TEXT, customer_id INTEGER, is_default INTEGER, is_active INTEGER)"
            )
        )
        s.commit()
        yield s


@pytest.fixture
def analysis_svc(session: Session) -> AnalysisService:
    return AnalysisService(session)


@pytest.fixture
def chemical_svc(session: Session) -> ChemicalService:
    return ChemicalService(session)


@pytest.fixture
def element_svc(session: Session) -> ElementService:
    return ElementService(session)


@pytest.fixture
def sample_parser(
    chemical_svc: ChemicalService,
    analysis_svc: AnalysisService,
    element_svc: ElementService,
) -> TRSampleFormParser:
    return TRSampleFormParser(
        chemical_service=chemical_svc,
        analysis_service=analysis_svc,
        element_service=element_svc,
        input_resolver=_StubResolver(),
    )


@pytest.fixture
def submission_parser(
    session: Session,
    chemical_svc: ChemicalService,
    analysis_svc: AnalysisService,
    element_svc: ElementService,
) -> TRSubmissionFormParser:
    return TRSubmissionFormParser(
        session=session,
        chemical_service=chemical_svc,
        analysis_service=analysis_svc,
        element_service=element_svc,
        input_resolver=_StubResolver(),
    )


# ---------------------------------------------------------------------------
# Helper: build a minimal Chemical worksheet
# ---------------------------------------------------------------------------

def _make_chemical_ws() -> Any:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Worksheet tab titles can't contain brackets; the service date comes
    # from the workbook *filename* which is passed separately to the parser.

    ws["D3"] = "Chemical Testing Request Form"
    ws["C10"] = "Acme Corp"
    ws["C11"] = "123 Main St"
    ws["C12"] = ""
    ws["I4"] = date(2026, 1, 15)
    ws["P5"] = date(2026, 1, 16)
    ws["C13"] = "John Smith"
    ws["B14"] = "Phone :"
    ws["C14"] = "555-1234"
    ws["P49"] = "Rev T"
    ws["H10"] = "PO-12345"
    ws["I11"] = "Visa 1234"
    ws["C15"] = "results@example.com"
    ws["C17"] = ""
    ws["N10"] = "invoice@example.com"
    ws["N11"] = ""

    # Header row = 20; put data in cols 3-15 to set last_col = 14 (B20 = "Sample ID" set after)
    for col in range(3, 16):
        ws.cell(row=20, column=col).value = f"H{col}"
    # Sample range: B20 = "Sample ID" → first_row = 22 (must be set after the loop above)
    ws["B20"] = "Sample ID"
    # One sample in row 22
    ws["B22"] = "S001"
    ws["C22"] = "Hydrochloric Acid"   # chemical name (col 3)
    ws["E22"] = "Si"                  # additional element (col 5)
    ws["K22"] = "Next Day"            # processing time (col 11)
    ws["M22"] = "2:00 PM"             # requested time (col 13)
    ws["N22"] = "Handle carefully"    # notes (col 14)
    # Footer: last row in col B that has data is row 33 (→ last_row = 33 - 11 = 22)
    ws["B33"] = "Total"

    return ws


# ---------------------------------------------------------------------------
# _parse_element_symbols
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw, expected",
    [
        ("Si,Fe,Cu", ["Si", "Fe", "Cu"]),
        ("Si;Fe;Cu", ["Si", "Fe", "Cu"]),
        ("Si Fe Cu", ["Si", "Fe", "Cu"]),
        ("Si.Fe.Cu", ["Si", "Fe", "Cu"]),
        ("Si, Fe , Cu", ["Si", "Fe", "Cu"]),
        ("", []),
        ("  ", []),
    ],
)
def test_parse_element_symbols(raw: str, expected: list[str]) -> None:
    assert _parse_element_symbols(raw) == expected


# ---------------------------------------------------------------------------
# _extract_service_date
# ---------------------------------------------------------------------------


def test_extract_service_date_iso_brackets() -> None:
    assert _extract_service_date("Lab [2026-01-20].xlsx") == date(2026, 1, 20)


def test_extract_service_date_underscore_separator() -> None:
    assert _extract_service_date("Lab [2026_01_20].xlsx") == date(2026, 1, 20)


def test_extract_service_date_no_brackets_returns_sentinel() -> None:
    assert _extract_service_date("Lab Results.xlsx") == SENTINEL_DATE


# ---------------------------------------------------------------------------
# _parse_emails
# ---------------------------------------------------------------------------


def test_parse_emails_single() -> None:
    assert _parse_emails("a@b.com") == ["a@b.com"]


def test_parse_emails_semicolon_separated() -> None:
    assert _parse_emails("a@b.com;c@d.com") == ["a@b.com", "c@d.com"]


def test_parse_emails_filters_invalid() -> None:
    assert _parse_emails("notanemail,a@b.com") == ["a@b.com"]


def test_parse_emails_empty_string() -> None:
    assert _parse_emails("") == []


# ---------------------------------------------------------------------------
# TRSampleFormParser
# ---------------------------------------------------------------------------


def test_sample_parser_builds_sample(
    sample_parser: TRSampleFormParser,
    session: Session,
) -> None:
    ws = _make_chemical_ws()
    bounds = SampleRangeBounds(first_row=22, last_row=22, first_col=2, last_col=14)
    sample = sample_parser.build_from_row(ws, row=22, sample_range=bounds, request_type=RequestType.CHEMICAL)

    assert sample.sample_name == "S001"
    assert sample.form_chemical_name == "Hydrochloric Acid"
    assert sample.chemical_id == _STUB_CHEMICAL.id
    assert sample.processing_time == ProcessingTime.NEXT_DAY
    assert sample.requested_time == "2:00 PM"
    assert sample.additional_notes == "Handle carefully"


def test_sample_parser_resolves_additional_elements(
    sample_parser: TRSampleFormParser,
) -> None:
    ws = _make_chemical_ws()
    bounds = SampleRangeBounds(first_row=22, last_row=22, first_col=2, last_col=14)
    sample = sample_parser.build_from_row(ws, row=22, sample_range=bounds, request_type=RequestType.CHEMICAL)
    # "Si" in col 5 should resolve to element_id=14
    assert 14 in sample.additional_element_ids


def test_sample_parser_analysis_ids_empty_when_form_analyses_stubbed(
    sample_parser: TRSampleFormParser,
) -> None:
    ws = _make_chemical_ws()
    bounds = SampleRangeBounds(first_row=22, last_row=22, first_col=2, last_col=14)
    sample = sample_parser.build_from_row(ws, row=22, sample_range=bounds, request_type=RequestType.CHEMICAL)
    assert sample.analysis_ids == ()


def test_sample_parser_water_uses_constant_chemical_name(
    sample_parser: TRSampleFormParser,
) -> None:
    ws = _make_chemical_ws()
    ws["D3"] = "Water Testing Request Form"
    ws["N22"] = "Next Day"   # WATER reads processing time from col 14
    bounds = SampleRangeBounds(first_row=22, last_row=22, first_col=2, last_col=14)
    sample = sample_parser.build_from_row(ws, row=22, sample_range=bounds, request_type=RequestType.WATER)
    assert sample.form_chemical_name == "Water"


# ---------------------------------------------------------------------------
# TRSubmissionFormParser
# ---------------------------------------------------------------------------


_CHEMICAL_FILENAME = "Lab Results [2026-01-20].xlsx"


def test_submission_parser_builds_submission(
    submission_parser: TRSubmissionFormParser,
) -> None:
    ws = _make_chemical_ws()
    fixed_dt = datetime(2026, 1, 16, 9, 0)
    sub = submission_parser.build_from_worksheet(ws, download_date=fixed_dt, filename=_CHEMICAL_FILENAME)

    assert sub.customer_id == _STUB_CUSTOMER.id
    assert sub.request_type == RequestType.CHEMICAL
    assert sub.date_submitted == date(2026, 1, 15)
    assert sub.date_received == date(2026, 1, 16)
    assert sub.customer_contact == "John Smith"
    assert sub.customer_phone == "555-1234"
    assert sub.po_information == "PO-12345"
    assert sub.download_date == fixed_dt


def test_submission_parser_attaches_samples(
    submission_parser: TRSubmissionFormParser,
) -> None:
    ws = _make_chemical_ws()
    sub = submission_parser.build_from_worksheet(ws, download_date=datetime.now(), filename=_CHEMICAL_FILENAME)
    assert len(sub.samples) == 1
    assert sub.samples[0].sample_name == "S001"


def test_submission_parser_reads_results_email(
    submission_parser: TRSubmissionFormParser,
) -> None:
    ws = _make_chemical_ws()
    sub = submission_parser.build_from_worksheet(ws, download_date=datetime.now(), filename=_CHEMICAL_FILENAME)
    assert "results@example.com" in sub.results_email_main


def test_submission_parser_service_date_from_filename(
    submission_parser: TRSubmissionFormParser,
) -> None:
    ws = _make_chemical_ws()
    sub = submission_parser.build_from_worksheet(ws, download_date=datetime.now(), filename=_CHEMICAL_FILENAME)
    assert sub.service_date == date(2026, 1, 20)


def test_submission_parser_form_customer_fields(
    submission_parser: TRSubmissionFormParser,
) -> None:
    ws = _make_chemical_ws()
    sub = submission_parser.build_from_worksheet(ws, download_date=datetime.now(), filename=_CHEMICAL_FILENAME)
    assert sub.form_customer_name == "Acme Corp"
    assert sub.form_customer_address == "123 Main St"
    assert sub.form_customer_address_2 == ""

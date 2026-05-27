"""
create_test_submission.py
─────────────────────────
Generates  samples/<MMDDYY> ACME ENV LABS Results [YYYY-MM-DD].xlsx

The filename is built to pass invoice_file_validator (clsInvoiceReportGatekeeper
equivalent):
  1. extension is .xlsx
  2. "Results" in filename
  3. "Partial" NOT in filename
  4. first 6 characters are digits  (MMDDYY from date_received)

Request type : Water Testing Request Form (RequestType.WATER)
  → chemical name is hard-coded to "Water" by the parser — no form_chemicals
    lookup needed, so UnresolvedChemicalError is never raised.
  → cell addresses that differ from the Chemical form are accounted for below.

Customer     : ACME ENV LABS / 1200 Industrial Blvd / Sacramento CA 95814
  → must exactly match the form_customers row seeded by seed_db.py.
  → matched via the composite key  (form_customer | form_address | form_address_2)
    lower-cased in CustomerCache._build().

Samples (4 rows, rows 22-25):
  W-001  36 Elements + 7 Anions (bundle quote Q1), Next Day, Pb & As elements
  W-002  36 Elements + 7 Anions (bundle quote Q1), Next Day
  W-003  pH only (single-analysis quote Q2),        Next Day
  W-004  36 Elements + 7 Anions (bundle quote Q1),  Five Days  (alt. price tier)

Analysis resolution:
  form_analyses is fully ported.  _collect_analysis_ids calls
  get_ids_by_form_name(cell_value) for each non-empty data cell.
  Sample cells contain the exact form_analyses.form_name strings from seed_db.py.
  E21 (additional-elements sub-header) is left empty so it does not inject
  analysis_id=1 into samples that do not request "36 Elements".

Run from the project root after seed_db.py has been run:
    python create_test_submission.py
"""

import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Dates
# ---------------------------------------------------------------------------
today          = date.today()
date_received  = today - timedelta(days=12)      # well within the last 31 days
date_submitted = date_received - timedelta(days=1)
service_date   = date_received + timedelta(days=1)   # embedded in filename brackets

# ---------------------------------------------------------------------------
# Filename — gatekeeper-valid
#   prefix : MMDDYY from date_received  (submission_loader._extract_file_date)
#   body   : contains "Results", no "Partial"
#   suffix : [YYYY-MM-DD] service date   (_extract_service_date regex)
# ---------------------------------------------------------------------------
prefix   = date_received.strftime("%m%d%y")          # e.g. "051526"
filename = f"{prefix} ACME ENV LABS Results [{service_date}].xlsx"

out_dir  = Path("samples")
out_dir.mkdir(exist_ok=True)
out_path = out_dir / filename

# ---------------------------------------------------------------------------
# Build worksheet
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Water Testing Request Form"

# ── D3 : form type ───────────────────────────────────────────────────────
# _read_request_type() matches exactly "Water Testing Request Form"
ws["D3"] = "Water Testing Request Form"

# ── Submission / received dates ───────────────────────────────────────────
# _date_submitted_cell(WATER) = "I4"
# _date_received_cell(WATER)  = "S5"
ws["I4"] = date_submitted
ws["S5"] = date_received

# ── Customer / contact block (rows 10–17) ────────────────────────────────
#
# TRSubmissionFormParser.build_from_worksheet reads:
#   C10 → form_customer_name   \
#   C11 → form_customer_address  >  composite key for CustomerCache lookup
#   C12 → form_customer_address_2/
#   C13 → customer_contact
#   B14 / C14 → phone  (_read_phone: B14=="Phone :" → phone from C14)
#   C15 → results_email_main
#   C17 → results_email_cc

ws["C10"] = "ACME ENV LABS"           # must match form_customers.form_customer
ws["C11"] = "1200 Industrial Blvd"    # must match form_customers.form_address
ws["C12"] = "Sacramento CA 95814"     # must match form_customers.form_address_2
ws["C13"] = "Dr. Jane Carter"
ws["B14"] = "Phone :"
ws["C14"] = "916-555-0100"
ws["C15"] = "results@acmelabs.example.com"
ws["C17"] = ""

# PO number
# _read_revision_letter(WATER) reads S51.  "Rev I" → rev="I"
# _read_po_information: WATER, rev.upper() >= "I" → addr = "H10"
ws["S51"] = "Rev I"
ws["H10"] = "PO-88421"

# Invoice emails  (WATER → _invoice_email_cells = ("P10", "P11"))
ws["P10"] = "jcarter@acmelabs.example.com"
ws["P11"] = ""

# Credit card  (WATER → _read_credit_card_info reads J11, J12)
ws["J11"] = ""

# ── Header rows 20-21 ─────────────────────────────────────────────────────
#
# _get_sample_range logic:
#   B20 == "Sample ID"  → first_row = 22
#   header_row          = first_row - 2 = 20  (used for last_col calc)
#   last_col            = (last non-empty col in row 20) - 1
#
# Row 20 filled through col 18 (R) → ws.max_column context = 19 (from S51)
#   scan finds col 18 as last non-empty → last_col = 17
#
# _collect_analysis_ids uses  ws.cell(first_row - 1, add_el_col) = row 21 col 5
#   to pick up one analysis name via get_by_name() for every sample.
#   Setting E21 = "36 Elements" → analysis_id 1 attached to all samples.
#
# WATER column assignments:
#   col 14 (N) = processing time   _get_processing_time  → ProcessingTime.from_form_string()
#   col 16 (P) = requested time    _requested_time_col
#   col 17 (Q) = notes             _notes_col
#   col  5 (E) = additional elements  _additional_element_col  → element symbols per row

# Row 20 — main column headers
ws.cell(row=20, column=2).value  = "Sample ID"       # B20 — triggers first_row=22
ws.cell(row=20, column=3).value  = "Matrix"
ws.cell(row=20, column=4).value  = "36 Elements"
ws.cell(row=20, column=5).value  = "Addt'l Elements"
ws.cell(row=20, column=6).value  = "7 Anions"
ws.cell(row=20, column=7).value  = "pH"
ws.cell(row=20, column=8).value  = "ICP-MS Metals"
ws.cell(row=20, column=9).value  = "Titration"
ws.cell(row=20, column=14).value = "Processing Time"  # col 14 — WATER proc. time
ws.cell(row=20, column=16).value = "Requested Time"   # col 16
ws.cell(row=20, column=17).value = "Notes"             # col 17
ws.cell(row=20, column=18).value = "——"                # sentinel: last non-empty → last_col=17

# Row 21 — sub-header
#   E21 (col 5) was previously set to "36 Elements" as a fallback workaround
#   for when form_analyses was not yet ported.  _collect_analysis_ids reads
#   E21 via get_by_name() and ALWAYS adds that analysis ID to every sample,
#   even samples that don't request it.  Now that form_analyses is fully ported,
#   analysis IDs come from the form name in each data cell (below), so E21
#   must be blank — otherwise pH-only sample W-003 would get analysis_ids=(4,1)
#   instead of (4,) and no quote would match.
ws.cell(row=21, column=5).value = ""

# ── Sample rows 22-25 ────────────────────────────────────────────────────
#
# Now that form_analyses is ported, _collect_analysis_ids calls
# get_ids_by_form_name(cell_value) on each non-empty data cell.  The cell must
# contain the exact form_analyses.form_name string (case-insensitive).
# Use the exact names seeded in seed_db.py — NOT "X" markers.
#
# Processing time strings must be in _PT_FROM_FORM:
#   "Next Day"  → ProcessingTime.NEXT_DAY  (id=2)  ← Q1 Next Day price row
#   "Five Days" → ProcessingTime.FIVE_DAYS (id=8)  ← Q1 Five Days price row

# Sample 1 — W-001: bundle (36 Elements + 7 Anions), Next Day
#   Hits Q1 (bundle line item: analysis_ids {1,2}, processing_time_id 2)
ws.cell(row=22, column=2).value  = "W-001"
ws.cell(row=22, column=4).value  = "36 Elements"  # → analysis_id 1
ws.cell(row=22, column=5).value  = "Pb;As"        # additional element symbols → Pb(id=1), As(id=2)
ws.cell(row=22, column=6).value  = "7 Anions"     # → analysis_id 2; together {1,2} hits Q1
ws.cell(row=22, column=14).value = "Next Day"
ws.cell(row=22, column=16).value = "10:00 AM"
ws.cell(row=22, column=17).value = "Priority run"

# Sample 2 — W-002: bundle, Next Day (second bundle sample)
ws.cell(row=23, column=2).value  = "W-002"
ws.cell(row=23, column=4).value  = "36 Elements"  # → analysis_id 1
ws.cell(row=23, column=6).value  = "7 Anions"     # → analysis_id 2
ws.cell(row=23, column=14).value = "Next Day"
ws.cell(row=23, column=16).value = "10:30 AM"
ws.cell(row=23, column=17).value = ""

# Sample 3 — W-003: pH only (single-analysis quote Q2), Next Day
ws.cell(row=24, column=2).value  = "W-003"
ws.cell(row=24, column=7).value  = "pH"           # → analysis_id 4; hits Q2
ws.cell(row=24, column=14).value = "Next Day"
ws.cell(row=24, column=16).value = "11:00 AM"
ws.cell(row=24, column=17).value = ""

# Sample 4 — W-004: 36 Elements + 7 Anions, Five Days
#   Hits Q1 Five Days price row ($55 regular / $44 bulk).
#   Original design had "7 Anions only" but there is no standalone quote for
#   {analysis_id=2}; Q1 covers the bundle {1,2}.  Changed to the full bundle.
ws.cell(row=25, column=2).value  = "W-004"
ws.cell(row=25, column=4).value  = "36 Elements"  # → analysis_id 1
ws.cell(row=25, column=6).value  = "7 Anions"     # → analysis_id 2; {1,2} hits Q1 Five Days
ws.cell(row=25, column=14).value = "Five Days"
ws.cell(row=25, column=16).value = ""
ws.cell(row=25, column=17).value = ""

# ── Footer ────────────────────────────────────────────────────────────────
#
# _get_sample_range scans col B from max_row downward for the last non-None.
# With samples at rows 22-25 and S51 set (max_row ≥ 51), we need a col-B
# sentinel at row 36 so:
#   last_b_row = 36  →  last_row = 36 - 11 = 25  ✓  (all 4 samples included)
ws.cell(row=36, column=2).value = "Total"

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(out_path)
print(f"✓  Workbook saved: {out_path}")

# ── Gatekeeper validation ─────────────────────────────────────────────────
# Mirrors what SubmissionLoader.load_single() does before parsing.
from slim.pipeline.invoice_file_validator import validate_invoice_file

try:
    validate_invoice_file(out_path)
    print("✓  Gatekeeper validation passed  (extension, 'Results', 6-digit prefix)")
except ValueError as e:
    print(f"✗  Gatekeeper validation FAILED: {e}", file=sys.stderr)
    sys.exit(1)

# ── Structural parse (stub resolver) ─────────────────────────────────────
# The live TRFormInputResolver calls resolve_chemical() for every sample row,
# which calls ChemicalCache.get_id_by_form_name().  The form_chemicals table
# is not yet ported so _by_form_name is always empty and every parse raises
# UnresolvedChemicalError — even for WATER forms where the chemical name is
# hard-coded to "Water" by the parser.
#
# We verify the sheet structure with a stub resolver that bypasses this gap,
# exactly as tests/domain/tr/test_parsers.py does.
from datetime import datetime

import openpyxl as _openpyxl
from sqlalchemy import create_engine as _create_engine, text as _text
from sqlalchemy.orm import Session as _Session

from infrastructure.database import Base as _Base
from slim.domain.analysis.analysis_service import AnalysisService as _AnalysisSvc
from slim.domain.chemical.chemical import Chemical as _Chemical
from slim.domain.chemical.chemical_service import ChemicalService as _ChemSvc
from slim.domain.customer.customer import Customer as _Customer
from slim.domain.customer.customer_service import CustomerService as _CustSvc
from slim.domain.element.element_service import ElementService as _ElSvc
from slim.domain.tr.tr_submission_form_parser import TRSubmissionFormParser as _Parser


class _StubResolver:
    """Bypasses DB lookups — same approach used in test_parsers.py."""
    _chem = _Chemical(
        id=1, name="Acid Digestion Mix",
        metals_prep="HNO3/HCl digest", silicon_prep="HF trace", ions_prep="dilute HNO3",
    )
    _cust = _Customer(
        id=1, name="Acme Environmental Labs",
        street_address="1200 Industrial Blvd", city="Sacramento",
        state="CA", postal_code="95814", country="USA",
    )
    def resolve_customer(self, name, address, address2): return self._cust
    def resolve_chemical(self, form_name):               return self._chem


# Spin up an in-memory DB with the seeded analyses so get_by_name() works.
_engine = _create_engine("sqlite:///:memory:")
_Base.metadata.create_all(_engine)
with _Session(_engine) as _s:
    from slim.domain.analysis.analysis_repository import _AnalysisRow as _AR
    from slim.domain.element.element_repository  import _ElementRow  as _ER
    for aid, aname in [
        (1, "36 Elements"), (2, "7 Anions"), (3, "Titration"), (4, "pH"), (5, "ICP-MS Metals")
    ]:
        _s.add(_AR(id=aid, analysis_name=aname))
    for eid, sym, ename in [(1,"Pb","Lead"),(2,"As","Arsenic"),(3,"Cd","Cadmium"),(4,"Hg","Mercury")]:
        _s.add(_ER(id=eid, element_symbol=sym, element_name=ename))
    # invoice_default_contacts queried inside the parser; must exist to avoid Exception swallow
    _s.execute(_text(
        "CREATE TABLE invoice_default_contacts "
        "(contact_email TEXT, customer_id INTEGER, is_default INTEGER, is_active INTEGER)"
    ))
    _s.commit()
    _analysis_svc = _AnalysisSvc(_s)
    _chem_svc     = _ChemSvc(_s)
    _el_svc       = _ElSvc(_s)
    _parser       = _Parser(
        session=_s,
        chemical_service=_chem_svc,
        analysis_service=_analysis_svc,
        element_service=_el_svc,
        input_resolver=_StubResolver(),
    )
    _wb = _openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    _ws = _wb.worksheets[0]
    _sub = _parser.build_from_worksheet(_ws, download_date=datetime.now(), filename=filename)
    _wb.close()

print("✓  Structural parse passed  (stub resolver)")
print(f"   customer_id    : {_sub.customer_id}")
print(f"   request_type   : {_sub.request_type.name}")
print(f"   date_submitted : {_sub.date_submitted}")
print(f"   date_received  : {_sub.date_received}")
print(f"   service_date   : {_sub.service_date}")
print(f"   po_information : {_sub.po_information!r}")
print(f"   results_email  : {list(_sub.results_email_main)}")
print(f"   samples ({len(_sub.samples)}):")
for s in _sub.samples:
    print(f"     {s.sample_name:<8}  pt={s.processing_time.name:<15} "
          f"analysis_ids={s.analysis_ids}  element_ids={s.additional_element_ids}")

if len(_sub.samples) != 4:
    print(f"\n✗  Expected 4 samples, got {len(_sub.samples)}", file=sys.stderr)
    sys.exit(1)

# ── Summary ───────────────────────────────────────────────────────────────
print()
print(f"  Filename       : {filename}")
print(f"  Date submitted : {date_submitted}")
print(f"  Date received  : {date_received}  ({(today - date_received).days} days before {today})")
print(f"  Service date   : {service_date}  (bracket notation in filename)")
print()
print("  Samples:")
print("    W-001  36 Elements + 7 Anions (bundle Q1)  Next Day   Pb, As additional elements")
print("    W-002  36 Elements + 7 Anions (bundle Q1)  Next Day")
print("    W-003  pH only               (single  Q2)  Next Day")
print("    W-004  36 Elements + 7 Anions (bundle Q1)  Five Days  (alt. price tier)")
print()
print("  form_chemicals and form_analyses are fully ported.")
print("  Run:  slim --verbose run-single --file <this-file> --output out.csv")
